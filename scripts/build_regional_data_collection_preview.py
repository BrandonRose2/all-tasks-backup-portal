#!/usr/bin/env python3
"""Build static browser previews for the Regional Data Collection task archive.

The source task stores Excel workbooks and email drafts as archival attachments.
GitHub Pages intentionally excludes the multi-gigabyte attachment directory, so this
script creates a small, targeted preview package under ``previews/`` and copies only
the final reviewed assets needed for the Regional Data Collection workspace.
"""

from __future__ import annotations

import html
import shutil
from pathlib import Path

from markdown import markdown
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
TASK_ID = "BC9veM4AKxKZRf8ketbDTP"
SOURCE_DIR = ROOT / "client" / "public" / "archives" / "manus" / "attachments" / TASK_ID
OUTPUT_DIR = ROOT / "previews" / "regional-data-collection"
FILES_DIR = OUTPUT_DIR / "files"

# These are the final reviewed assets in the archived task's chronological sequence.
ASSETS = {
    "email_drafts_preview.md": "57-assistant_message-0_5d5e3053f09d786d_email_drafts_preview.md",
    "email_bodies_copy_paste.md": "117-assistant_message-0_8cb8f2fd80e88705_email_bodies_copy_paste.md",
    "group_email_draft.md": "183-assistant_message-0_92d4c6941f90922f_group_email_draft.md",
    "Region_1_JR_Rolon_Data_Collection.xlsx": "117-assistant_message-1_955b781069859c01_Region_1_JR_Rolon_Data_Collection.xlsx",
    "Region_2_Susan_Lopez_Data_Collection.xlsx": "117-assistant_message-2_b8ee5d9ef1d19e50_Region_2_Susan_Lopez_Data_Collection.xlsx",
    "Region_3_Ginger_Positerry_Data_Collection.xlsx": "117-assistant_message-3_f888a6c30b6836f5_Region_3_Ginger_Positerry_Data_Collection.xlsx",
    "Region_4_Blake_Weddington_Data_Collection.xlsx": "117-assistant_message-4_d5e002bdf651ca45_Region_4_Blake_Weddington_Data_Collection.xlsx",
    "Region_5_Leslie_Rolon_Data_Collection.xlsx": "117-assistant_message-5_c0c7f4cbce752946_Region_5_Leslie_Rolon_Data_Collection.xlsx",
    "Consolidated_All_Regions_Data_Collection.xlsx": "130-assistant_message-0_b54fe084d704b3a4_Consolidated_All_Regions_Data_Collection.xlsx",
    "ApartmentCorp_Shared_Data_Collection.xlsx": "166-assistant_message-0_ccf1fb5439686513_ApartmentCorp_Shared_Data_Collection.xlsx",
}

WORKBOOKS = [
    ("Region 1 — JR Rolon", "Region_1_JR_Rolon_Data_Collection.xlsx"),
    ("Region 2 — Susan Lopez", "Region_2_Susan_Lopez_Data_Collection.xlsx"),
    ("Region 3 — Ginger Positerry", "Region_3_Ginger_Positerry_Data_Collection.xlsx"),
    ("Region 4 — Blake Weddington", "Region_4_Blake_Weddington_Data_Collection.xlsx"),
    ("Region 5 — Leslie Rolon", "Region_5_Leslie_Rolon_Data_Collection.xlsx"),
    ("Consolidated all regions", "Consolidated_All_Regions_Data_Collection.xlsx"),
    ("Shared collection workbook", "ApartmentCorp_Shared_Data_Collection.xlsx"),
]

STYLE = """
:root { color-scheme: dark; --bg:#09101c; --card:#111b2c; --line:rgba(172,194,226,.18); --ink:#eef4ff; --muted:#a2b2c9; --accent:#90b9ff; --accent2:#5eead4; }
*{box-sizing:border-box} body{margin:0;background:radial-gradient(circle at 10% -10%,rgba(37,99,235,.26),transparent 32rem),var(--bg);color:var(--ink);font:15px/1.55 Inter,system-ui,sans-serif} a{color:var(--accent)} .wrap{width:min(1280px,calc(100% - 32px));margin:0 auto;padding:32px 0 52px} .crumb{color:var(--accent);font-size:12px;font-weight:800;letter-spacing:.09em;text-transform:uppercase;text-decoration:none}.hero{margin:18px 0 26px}.hero h1{font-size:clamp(29px,5vw,48px);letter-spacing:-.045em;line-height:1.05;margin:0}.hero p{max-width:790px;color:var(--muted);margin:13px 0 0}.notice{margin:20px 0;padding:12px 14px;border:1px solid rgba(251,191,36,.3);border-radius:12px;background:rgba(251,191,36,.08);color:#fde68a;font-size:13px}.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(235px,1fr));gap:12px}.card{display:flex;min-height:150px;flex-direction:column;justify-content:space-between;border:1px solid var(--line);border-radius:16px;padding:16px;color:var(--ink);text-decoration:none;background:linear-gradient(145deg,rgba(22,33,52,.95),rgba(13,22,36,.92));transition:transform .15s,border-color .15s}.card:hover{transform:translateY(-2px);border-color:var(--accent)}.kind{color:var(--accent2);font-size:10px;font-weight:800;letter-spacing:.09em}.card h2{font-size:16px;line-height:1.3;margin:12px 0 4px}.card p{margin:0;color:var(--muted);font-size:12px}.section{margin-top:34px}.section h2{font-size:21px;letter-spacing:-.025em;margin:0 0 10px}.copy{border:1px solid var(--line);border-radius:16px;background:rgba(17,27,44,.82);padding:22px;overflow:auto}.copy h1,.copy h2,.copy h3{line-height:1.2}.copy h1{font-size:27px}.copy h2{font-size:20px;margin-top:30px}.copy blockquote{margin:14px 0;padding:0 14px;border-left:3px solid var(--accent);color:var(--muted)}.copy table{width:100%;border-collapse:collapse;overflow:hidden;margin:15px 0}.copy th,.copy td{padding:8px;border:1px solid var(--line);text-align:left;vertical-align:top}.copy th{background:rgba(96,165,250,.13)}.toolbar{display:flex;gap:10px;flex-wrap:wrap;margin:18px 0}.button{padding:9px 12px;border:1px solid var(--line);border-radius:10px;background:var(--card);color:var(--ink);text-decoration:none;font-size:12px;font-weight:700}.button:hover{border-color:var(--accent);color:var(--accent)}.sheet{margin-top:20px;border:1px solid var(--line);border-radius:14px;overflow:auto;background:rgba(17,27,44,.86)}.sheet h2{position:sticky;left:0;margin:0;padding:12px 14px;font-size:16px;background:#152139;border-bottom:1px solid var(--line)}table.workbook{border-collapse:collapse;min-width:760px;width:max-content}table.workbook td,table.workbook th{border:1px solid var(--line);padding:6px 8px;min-width:90px;max-width:340px;white-space:pre-wrap;vertical-align:top;font-size:12px}table.workbook th{position:sticky;top:0;background:#17243d;color:#cbd8ec}table.workbook tr:nth-child(even) td{background:rgba(255,255,255,.015)}.foot{margin-top:32px;color:var(--muted);font-size:11px}@media(max-width:600px){.wrap{width:min(100% - 20px,1280px);padding-top:22px}.copy{padding:16px}}
"""


def page(title: str, body: str) -> str:
    return f"""<!doctype html><html lang=\"en\"><head><meta charset=\"utf-8\"><meta name=\"viewport\" content=\"width=device-width,initial-scale=1\"><meta name=\"robots\" content=\"noindex,nofollow\"><title>{html.escape(title)}</title><style>{STYLE}</style></head><body><main class=\"wrap\">{body}<footer class=\"foot\">Regional Data Collection preview · Source task archive preserved in the public backup repository. This static preview is for review only and does not send email or modify any source workbook.</footer></main></body></html>"""


def copy_assets() -> None:
    if OUTPUT_DIR.exists():
        shutil.rmtree(OUTPUT_DIR)
    FILES_DIR.mkdir(parents=True, exist_ok=True)
    for display_name, source_name in ASSETS.items():
        source = SOURCE_DIR / source_name
        if not source.exists():
            raise FileNotFoundError(f"Missing archived source asset: {source}")
        shutil.copy2(source, FILES_DIR / display_name)


def preview_workbook(display_name: str) -> str:
    input_path = FILES_DIR / display_name
    workbook = load_workbook(input_path, read_only=True, data_only=False)
    sheets: list[str] = []
    for worksheet in workbook.worksheets:
        max_rows = min(worksheet.max_row or 1, 120)
        max_columns = min(worksheet.max_column or 1, 32)
        header_cells = "".join(f"<th>{html.escape(str(col))}</th>" for col in range(1, max_columns + 1))
        rows: list[str] = []
        for row_number, row in enumerate(worksheet.iter_rows(min_row=1, max_row=max_rows, max_col=max_columns, values_only=True), start=1):
            cells = "".join(f"<td>{html.escape('' if value is None else str(value))}</td>" for value in row)
            rows.append(f"<tr><th>{row_number}</th>{cells}</tr>")
        sheets.append(f"<section class=\"sheet\"><h2>{html.escape(worksheet.title)}</h2><table class=\"workbook\"><thead><tr><th>Row</th>{header_cells}</tr></thead><tbody>{''.join(rows)}</tbody></table></section>")
    workbook.close()
    label = display_name.removesuffix(".xlsx").replace("_", " ")
    return page(
        f"{label} — Preview",
        f"<a class=\"crumb\" href=\"index.html\">← Regional Data Collection</a><header class=\"hero\"><h1>{html.escape(label)}</h1><p>Browser-readable spreadsheet preview. Formula cells are shown as saved formulas; formatting and data validation are retained in the downloadable Excel original.</p></header><div class=\"toolbar\"><a class=\"button\" href=\"files/{html.escape(display_name)}\" download>Download original Excel workbook</a><a class=\"button\" href=\"https://github.com/BrandonRose2/all-tasks-backup-portal/tree/main/client/public/archives/manus/attachments/{TASK_ID}\" target=\"_blank\" rel=\"noopener\">View full attachment archive</a></div>{''.join(sheets)}",
    )


def main() -> None:
    copy_assets()
    email_preview = (FILES_DIR / "email_drafts_preview.md").read_text(encoding="utf-8")
    markdown_html = markdown(email_preview, extensions=["tables", "fenced_code", "sane_lists"])

    for label, display_name in WORKBOOKS:
        (OUTPUT_DIR / f"{display_name.removesuffix('.xlsx')}.html").write_text(preview_workbook(display_name), encoding="utf-8")

    workbook_cards = "".join(
        f"<a class=\"card\" href=\"{html.escape(display_name.removesuffix('.xlsx'))}.html\"><span class=\"kind\">EXCEL PREVIEW</span><div><h2>{html.escape(label)}</h2><p>Open a browser-readable preview or download the original workbook.</p></div></a>"
        for label, display_name in WORKBOOKS
    )
    index_body = f"""
    <a class=\"crumb\" href=\"../../index.html\">← Task Backup Portal</a>
    <header class=\"hero\"><h1>Regional Data Collection</h1><p>Review workspace for the archived Manager’s Data Collection task. It includes final email drafts and browser-readable previews of the preserved regional Excel templates.</p></header>
    <div class=\"notice\"><strong>Review-only:</strong> opening any preview does not send email or change the original workbook. Download buttons retrieve the preserved Excel originals.</div>
    <div class=\"toolbar\"><a class=\"button\" href=\"files/email_drafts_preview.md\" download>Download email-draft markdown</a><a class=\"button\" href=\"files/email_bodies_copy_paste.md\" download>Download copy/paste bodies</a><a class=\"button\" href=\"files/group_email_draft.md\" download>Download group email draft</a><a class=\"button\" href=\"https://github.com/BrandonRose2/all-tasks-backup-portal/tree/main/client/public/archives/manus/attachments/{TASK_ID}\" target=\"_blank\" rel=\"noopener\">View full archived attachments</a></div>
    <section class=\"section\"><h2>Email drafts preview</h2><div class=\"copy\">{markdown_html}</div></section>
    <section class=\"section\"><h2>Workbook previews</h2><div class=\"grid\">{workbook_cards}</div></section>
    """
    (OUTPUT_DIR / "index.html").write_text(page("Regional Data Collection — Preview", index_body), encoding="utf-8")
    (OUTPUT_DIR / "README.md").write_text(
        "# Regional Data Collection Preview\n\nGenerated from archived task `BC9veM4AKxKZRf8ketbDTP` by `scripts/build_regional_data_collection_preview.py`.\n",
        encoding="utf-8",
    )
    print(f"Built Regional Data Collection preview at {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
